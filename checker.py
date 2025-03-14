from enum import Enum
import os
import sys
import json
import io
import re
import importlib.util
import unittest
from unittest.mock import patch
from dotenv import load_dotenv
import requests
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from utils.consignas import EJOB_1_B

UTEST_DIR = "tests"
UTEST_FILENAME = "ejB.test.py"
EXAMS_DIR = "parciales"
ASSIGNMENT = EJOB_1_B


load_dotenv()


class TestResult(Enum):
    PASSED = "PASSED"
    FAILED = "FAILED"
    ERROR = "ERROR"


class JSONTestResult(unittest.TestResult):
    """A custom TestResult class that collects test results in JSON format.

    This class extends unittest.TestResult to store test execution results in a structured JSON format.
    Each test result includes the test name, description, result status (PASSED, FAILED, ERROR),
    and a reason or error message.

    Attributes:
      results (list): A list of dictionaries containing the test results.
      result_options (TestResult): An enum containing the possible test result statuses.

    Methods:
      addSuccess(test): Records a successful test execution.
      addFailure(test, err): Records a test failure.
      addError(test, err): Records a test error.

    Each result dictionary contains:
      - test: The name of the test method
      - description: The test's description (from docstring)
      - result: The test result status (PASSED/FAILED/ERROR)
      - reason: Explanation of the result or error message
    """

    def __init__(self, *args, **kwargs):
        super(JSONTestResult, self).__init__(*args, **kwargs)
        self.results = []
        self.result_options = TestResult

    def addSuccess(self, test):
        description = test.shortDescription() or ""
        self.results.append({
            "test": test._testMethodName,
            "description": description,
            "result": self.result_options.PASSED.value,
            "reason": "Test passed successfully"
        })
        super(JSONTestResult, self).addSuccess(test)

    def addFailure(self, test, err):
        description = test.shortDescription() or ""
        reason = self._exc_info_to_string(err, test)
        self.results.append({
            "test": test._testMethodName,
            "description": description,
            "result": self.result_options.FAILED.value,
            "reason": reason
        })
        super(JSONTestResult, self).addFailure(test, err)

    def addError(self, test, err):
        description = test.shortDescription() or ""
        reason = self._exc_info_to_string(err, test)
        self.results.append({
            "test": test._testMethodName,
            "description": description,
            "result": self.result_options.ERROR.value,
            "reason": reason
        })
        super(JSONTestResult, self).addError(test, err)


def load_module_from_path(module_name, module_path):
    """
    Loads a Python module from a given file path.

    This function dynamically imports a Python module from a specified file path using the importlib
    utility. It creates a module specification, loads the module, and executes it.

    Args:
      module_name (str): The name to be given to the loaded module
      module_path (str): The file path where the module is located

    Returns:
      module: The loaded Python module object

    Raises:
      ImportError: If the module cannot be loaded
      AttributeError: If the module does not contain expected attributes
    """
    spec = importlib.util.spec_from_file_location(module_name, module_path)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def load_test_module_from_path(module_name, module_path):
    """
    Dynamically loads a test module from a given file path with special handling for exam function imports.
    This function reads a Python test file and modifies how exam functions are imported by wrapping
    each import in a try/except block. This allows for graceful handling of missing functions.
    Args:
      module_name (str): The name to give to the loaded module
      module_path (str): The file path to the test module to load
    Returns:
      module: The loaded test module object with modified imports
    Raises:
      FileNotFoundError: If the specified module_path does not exist
      SyntaxError: If the source code in the file has syntax errors
    Example:
      >>> test_module = load_test_module_from_path("test_exam", "./test_exam.py")
      >>> test_module.some_test_function()
    Note:
      The function specifically looks for and modifies imports in the format:
      'from exam import function1, function2, ...'
    """
    with open(module_path, "r") as f:
        test_source = f.read()
    pattern = r"^from\s+exam\s+import\s+(.+)$"
    match = re.search(pattern, test_source, re.MULTILINE)
    if match:
        imported_functions = match.group(1)
        func_list = [fn.strip() for fn in imported_functions.split(",")]
        #!Generate try/except blocks for each function import
        replacement = ""
        for fn in func_list:
            replacement += f"try:\n    from exam import {fn}\n"
            replacement += f"except ImportError as e:\n"
            replacement += f"    def {fn}(*args, **kwargs):\n"
            replacement += f"        raise ImportError('Function \"{fn}\" not found. ' + str(e))\n"
        test_source = re.sub(pattern, replacement,
                             test_source, count=1, flags=re.MULTILINE)
    spec = importlib.util.spec_from_loader(module_name, loader=None)
    test_module = importlib.util.module_from_spec(spec)
    exec(test_source, test_module.__dict__)
    return test_module


def run_tests_for_student(student_module_path, test_module_path, assignment):
    """
    Runs unit tests for a student's module against a test module and returns the test results.

    This function attempts to load a student's Python module and a corresponding test module,
    then executes all tests found in the test module against the student's code. The function

    Parameters:
    ----------
    student_module_path : str
      File path to the student's Python module to be tested
    test_module_path : str
      File path to the test module containing the unit tests

    Returns:
    -------
    list
      A list of dictionaries containing test results. Each dictionary has the following keys:
      - 'test': str, name of the test or error type
      - 'description': str, description of the test or error
      - 'result': str, test outcome ('passed' or 'failed')
      - 'reason': str, explanation of failure (if applicable)

    Possible Error Returns:
    --------------------
    - ModuleLoadError: When student's module fails to load
    - TestModuleLoadError: When test module fails to load
    - TestExecutionError: When an error occurs during test execution

    Notes:
    -----
    The function uses JSONTestResult class for formatting test results
    and temporarily registers the student's module in sys.modules as "exam" to allow proper importing.
    """
    try:
        student_module = load_module_from_path("exam", student_module_path)
        # ! makes  "from exam import ..." work in test_module
        sys.modules["exam"] = student_module
    except Exception as e:
        return [{
            "test": "ModuleLoadError",
            "description": "El módulo del estudiante no pudo cargarse",
            "result": "failed",
            "reason": str(e)
        }]
    try:
        #! Load modify test module
        test_module = load_test_module_from_path(
            "exam_test_module", test_module_path)
    except Exception as e:
        return [{
            "test": "TestModuleLoadError",
            "description": "El módulo de tests no pudo cargarse",
            "result": "failed",
            "reason": str(e)
        }]
    try:
        #! run all tests
        suite = unittest.defaultTestLoader.loadTestsFromModule(test_module)
        stream = io.StringIO()
        runner = unittest.TextTestRunner(
            stream=stream, resultclass=JSONTestResult, verbosity=2)
        result = runner.run(suite)

        # Read student's source code
        with open(student_module_path, 'r') as f:
            student_code = f.read()

        # Read test source code
        with open(test_module_path, 'r') as f:
            test_code = f.read()

        # Add source codes to results
        return {
            "results": result.results,
            "exam": student_code,
            "testCode": test_code,
            "assignment": assignment
        }
    except Exception as e:
        return [{
            "test": "TestExecutionError",
            "description": "Error durante la ejecución de los tests",
            "result": "failed",
            "reason": str(e)
        }]


def main_runner(exams_dir, utest_dir, utest_filename, assignment) -> str:
    """
    Executes unit tests for Python scripts in a specified directory and collects results.
    This function processes Python files in the EXAMS_DIR directory, running unit tests
    defined in UTEST_DIR/UTEST_FILENAME against each student's submission. Results are
    collected and returned as a JSON formatted output.
    Returns:
      JSON-formatted dictionary where:
        - Keys are student names (derived from filenames without extension)
        - Values are test results for each student's submission
    Example of output format:
      {
        "student1": {test_results},
        "student2": {test_results},
        ...
      }
    Notes:
      - Assumes student's name is the filename without extension
      - Expects test module in UTEST_DIR/UTEST_FILENAME
      - Skips non-Python files in EXAMS_DIR
    """
    final_results = {}
    test_module_path = os.path.join(utest_dir, utest_filename)

    if not os.path.isdir(exams_dir):
        print("Exams directory not found")
        return ""

    #! Loop over exams files - requires format studentname.py
    python_files = [file for file in os.listdir(
        exams_dir) if file.endswith(".py")]

    print(f"Found {len(python_files)} valid python files...")
    for i, file in enumerate(python_files, 1):
        student_module_path = os.path.join(exams_dir, file)
        student_name = os.path.splitext(file)[0]
        print(
            f"[{i}/{len(python_files)}] Running tests for student: {student_name}", file=sys.stderr)
        test_output = run_tests_for_student(
            student_module_path, test_module_path, assignment)
        print(f"✓ Completed tests for: {student_name}", file=sys.stderr)

        if isinstance(test_output, list):  # Error case
            final_results[student_name] = {
                "results": test_output,
                "exam": "",
                "testCode": "",
                "assignment": assignment
            }
        else:  # Success case
            final_results[student_name] = test_output
    print("All tests completed.")
    return json.dumps(final_results, indent=4)


if __name__ == "__main__":
    os.makedirs("informes", exist_ok=True)
    json_results = main_runner(exams_dir=EXAMS_DIR, utest_dir=UTEST_DIR,
                               utest_filename=UTEST_FILENAME, assignment=ASSIGNMENT)

    try:
        print("Waiting AI response...")
        parsed_json = json.loads(json_results)

        #!AI
        webhook_url = os.getenv("WEBHOOK_URL")
        webhook_secret = os.getenv("WEBHOOK_SECRET")
        response = requests.post(
            webhook_url,
            json=parsed_json,
            headers={
                'CheckerCredentials': webhook_secret,
                'Content-Type': 'application/json',
            }
        )

        response.raise_for_status()

        response_data = response.json()

        # Process each item in the response data
        if "data" in response_data:
            for item in response_data["data"]:
                if "exam" in item:
                    item["exam"] = item["exam"].encode().decode(
                        "unicode_escape")
                if "testCode" in item:
                    item["testCode"] = item["testCode"].encode().decode(
                        "unicode_escape")

                for test in item.get("results", []):
                    if "reason" in test:
                        test["reason"] = test["reason"].encode().decode(
                            "unicode_escape")

        #!Excel report
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_filename = f"informes/results_{timestamp}.xlsx"
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Resultados de Exámenes"

            # Define styles
            header_fill = PatternFill(
                start_color="4472C4", end_color="4472C4", fill_type="solid")
            passed_fill = PatternFill(
                start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            failed_fill = PatternFill(
                start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            error_fill = PatternFill(
                start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            section_fill = PatternFill(
                start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")

            header_font = Font(bold=True, color="FFFFFF")

            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Main headers to repeat for each student
            main_headers = ["Nombre del Alumno", "Resultado Global",
                            "Nota Sugerida", "Confianza", "Informe"]

            # Set column widths
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 80

            # Helper function to add headers
            def add_main_headers(row_num):
                for col_num, header in enumerate(main_headers, 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.value = header
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(
                        horizontal='center', vertical='center', wrap_text=True)
                    cell.border = border
                return row_num + 1

            # Add title at the top of the document
            title_row = 1
            ws.merge_cells(start_row=title_row, start_column=1,
                           end_row=title_row, end_column=5)
            title_cell = ws.cell(row=title_row, column=1)
            title_cell.value = "REPORTE DE EVALUACIONES"
            title_cell.font = Font(bold=True, size=14)
            title_cell.alignment = Alignment(horizontal='center')

            # Starting row for first student
            row_num = 3

            if "data" in response_data:
                for item in response_data["data"]:
                    if "name" in item:
                        # Add separator between students
                        if row_num > 3:  # Not the first student
                            ws.merge_cells(
                                start_row=row_num, start_column=1, end_row=row_num, end_column=5)
                            separator = ws.cell(row=row_num, column=1)
                            separator.value = "───────────────────────────────────────────────────────────────"
                            separator.alignment = Alignment(
                                horizontal='center')
                            row_num += 1

                        # Add student section title
                        ws.merge_cells(
                            start_row=row_num, start_column=1, end_row=row_num, end_column=5)
                        student_section = ws.cell(row=row_num, column=1)
                        student_section.value = f"Estudiante: {item['name']}"
                        student_section.font = Font(bold=True, size=12)
                        student_section.fill = section_fill
                        student_section.alignment = Alignment(
                            horizontal='center')
                        row_num += 1

                        # Add main headers for this student
                        row_num = add_main_headers(row_num)

                        # Add student data
                        ws.cell(row=row_num, column=1).value = item["name"]

                        # Calculate test results
                        if "testResults" in item:
                            total_tests = len(item["testResults"])
                            passed_tests = sum(
                                1 for test in item["testResults"] if test["result"] == "PASSED")
                            result_cell = ws.cell(row=row_num, column=2)
                            result_cell.value = f"{passed_tests}/{total_tests} ({passed_tests * 100 // total_tests}%)"

                            if passed_tests == total_tests:
                                result_cell.fill = passed_fill
                            elif passed_tests >= total_tests // 2:
                                result_cell.fill = PatternFill(
                                    start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                            else:
                                result_cell.fill = failed_fill

                        # Add output data
                        if "output" in item:
                            if "nota_sugerida" in item["output"]:
                                nota_cell = ws.cell(row=row_num, column=3)
                                nota_cell.value = item["output"]["nota_sugerida"]

                                # Color based on grade
                                if item["output"]["nota_sugerida"] >= 70:
                                    nota_cell.fill = passed_fill
                                elif item["output"]["nota_sugerida"] >= 60:
                                    nota_cell.fill = PatternFill(
                                        start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                                else:
                                    nota_cell.fill = failed_fill

                            if "confianza" in item["output"]:
                                ws.cell(
                                    row=row_num, column=4).value = item["output"]["confianza"]

                            if "informe" in item["output"]:
                                ws.cell(
                                    row=row_num, column=5).value = item["output"]["informe"]
                                ws.cell(row=row_num, column=5).alignment = Alignment(
                                    wrap_text=True, vertical='top')

                        # Add borders to all cells in the student data row
                        for col in range(1, 6):
                            ws.cell(row=row_num, column=col).border = border

                        row_num += 2

                        # Add detailed test results title
                        ws.merge_cells(
                            start_row=row_num, start_column=1, end_row=row_num, end_column=5)
                        test_header = ws.cell(row=row_num, column=1)
                        test_header.value = f"Resultados detallados de pruebas"
                        test_header.font = Font(bold=True)
                        test_header.fill = section_fill
                        test_header.alignment = Alignment(horizontal='center')
                        row_num += 1

                        # Test results table headers
                        test_headers = [
                            "Prueba", "Descripción", "Resultado", "Razón"]
                        for col_num, header in enumerate(test_headers, 1):
                            cell = ws.cell(row=row_num, column=col_num)
                            cell.value = header
                            cell.font = Font(bold=True)
                            cell.fill = section_fill
                            cell.alignment = Alignment(horizontal='center')
                            cell.border = border
                        row_num += 1

                        # Add test results
                        if "testResults" in item:
                            for test in item["testResults"]:
                                ws.cell(row=row_num, column=1).value = test.get(
                                    "test", "")
                                ws.cell(row=row_num, column=2).value = test.get(
                                    "description", "")

                                result_cell = ws.cell(row=row_num, column=3)
                                result_cell.value = test.get("result", "")
                                if test.get("result") == "PASSED":
                                    result_cell.fill = passed_fill
                                elif test.get("result") == "FAILED":
                                    result_cell.fill = failed_fill
                                else:
                                    result_cell.fill = error_fill

                                ws.cell(row=row_num, column=4).value = test.get(
                                    "reason", "")
                                ws.cell(row=row_num, column=4).alignment = Alignment(
                                    wrap_text=True)

                                # Add borders to all cells in this row
                                for col in range(1, 5):
                                    ws.cell(row=row_num,
                                            column=col).border = border

                                row_num += 1

                        # Add spacing between students
                        row_num += 2

            # Save the workbook
            wb.save(excel_filename)
            print(f"Results saved to:")
            print(f"- Excel: {excel_filename}")
        except json.JSONDecodeError as e:
            print(f"Error: Invalid JSON format: {e}")
    except requests.exceptions.RequestException as e:
        print(f"Error making request to webhook: {e}")
    except Exception as e:
        print(f"Error saving results: {e}")
